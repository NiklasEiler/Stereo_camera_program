import bildverarbeitung_seg_unet as bv
import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
import cv2
import re
import numpy as np
import shutil
import pickle  
import tensorflow
import pandas as pd
import openpyxl
import open3d as o3d
from sklearn.cluster import DBSCAN
from sklearn.cluster import KMeans
from collections import Counter
from sklearn.cluster import DBSCAN
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pywinauto.application import Application
import time
import sys


def adjust_column_width(source_file):
    # Launch Excel
    app = Application().start(source_file)

    # Wait for Excel to open
    time.sleep(5)

    # Connect to the Excel window
    window = app.top_window()

    # Press Alt, then H (for Home tab)
    window.type_keys("%h")
    time.sleep(1)

    # Press O (for Format dropdown)
    window.type_keys("o")
    time.sleep(1)

    # Press I (for AutoFit Column Width)
    window.type_keys("i")
    time.sleep(1)

    print("Column widths adjusted successfully!")
    app.kill()

#load lin 
def load_modells():                
    #random forrest
    with open('Modelle/random_forest_model.pkl', 'rb') as file:
        rfreg = pickle.load(file)

    #load svm
    with open('Modelle/svm_pos_model.pkl', 'rb') as file:
        svm_pos = pickle.load(file)


    with open('Modelle/scalar_model.pkl', 'rb') as file:
        scaler = pickle.load(file)

    return rfreg, svm_pos, scaler

def natural_sort_key(s):
    # Regular expression to split the string into numeric and non-numeric parts
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]

def get_folder_names(path):
    folder_names = []
    for entry in os.listdir(path):
        if os.path.isdir(os.path.join(path, entry)):
            folder_names.append(entry)
    return folder_names


def trans_to_klasse_back(pred):
    pred[pred[:, 0] <= 25, 0] = 0
    pred[(pred[:, 0] > 25) & (pred[:, 0] < 75), 0] = 50
    pred[pred[:, 0] >= 75, 0] = 100

    pred[pred[:, 1] < 50, 1] = 0
    pred[pred[:, 1] >= 50, 1] = 100

    pred[pred[:, 2] < 50, 2] = 0
    pred[pred[:, 2] >= 50, 2] = 100

    pred=pred.astype(str)

    pred[pred[:, 0] == '0.0', 0] = 'schräg'
    pred[pred[:, 0] == '50.0', 0] = 'rechts'
    pred[pred[:, 0] == '100.0', 0] = 'mittig'

    pred[pred[:, 1] == '0.0', 1] = 'korrekt'
    pred[pred[:, 1] == '100.0', 1] = 'falsch'

    pred[pred[:, 2] == '0.0', 2] = 'bad'
    pred[pred[:, 2] == '100.0', 2] = 'good'

    return pred

def most_used_string(row):
    return row.mode()[0]

def get_sheet_names(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names

def image_and_data(image_folder, save_path, speichern_data, foldername, x=-1, y=-1):
    data_recieved=False
    labels= ['Dateiname','A','max Grad', 'min Grad', 'Var Grad', 'mittel Grad', 'median Grad', 'modus Grad', 'Rundheit', 'Umfang', 'Bauteil koor. x', 'Bauteil koor. y', 'Flächenmitte koor. x', 'Flächenmitte koor. y', 'Weißerstrich koor. x', 'Weißerstrich koor. y','delta FB','delta WB','delta FW', 'Winkel Weiß zu Max', 'd1', 'd2', 'd3']
    i=0
    filenames=[]
    size=len(os.listdir(image_folder))
    data=np.zeros((size,23)) 
    index=np.arange(1,size+1,1)

    for filename in sorted(os.listdir(image_folder), key=natural_sort_key): 
        
        if filename.endswith('.jpg') or filename.endswith('.png'):
            image_path = os.path.join(image_folder, filename)
            image = cv2.imread(image_path)

            if image is not None:
                img_save_path=save_path + "/" + filename
                print('###################################' + str(filename)+ '########################################')
                data[i,1:23] = bv.image_it(x, y, i,data, image ,img_save_path, vis=False, speicher_img=True, speichern_data=True)
                i=i+1
                data_recieved=True
                filenames.append(filename)

    if speichern_data and data_recieved:
        bv.speichern_csv_excel_fkt(data, labels,index, foldername, filenames)
    

def predection_daten_images(folder_names, data_path):
    columns_names = ['Position', 'Werkzeugeinbau', 'Ausschuss oder gut Teil', 'Position', 'Werkzeugeinbau', 'Ausschuss oder gut Teil', 'Position', 'Werkzeugeinbau', 'Ausschuss oder gut Teil', 'Position', 'Werkzeugeinbau', 'Ausschuss oder gut Teil', 'Position', 'Werkzeugeinbau', 'Ausschuss oder gut Teil']
    rfreg, svm_pos, scaler = load_modells()


    
    excel_sheet_names= get_sheet_names(data_path)
    
    for folder_name in folder_names:
        filenames_predict=[]

        if folder_name in excel_sheet_names:
            file_path = data_path
            # Load the Excel file into a Pandas DataFrame
            daten = pd.read_excel(file_path,  sheet_name=folder_name)
            
            daten.index = range(1, len(daten) + 1)
            filenames_predict=daten['Dateiname']
            daten.drop('Dateiname', axis=1, inplace=True)
            daten.drop('Unnamed: 0', axis=1, inplace=True)

            #normalise
            daten_scaled = scaler.transform(daten)
            daten_scaled = pd.DataFrame(daten_scaled,columns=daten.columns)

            #prediction
            pred_y_rf = rfreg.predict(daten_scaled)
            pred_y_rf= trans_to_klasse_back(pred_y_rf)

            pred_y_svm_pos = svm_pos.predict(daten_scaled)
            pred_y_svm_pos=pred_y_svm_pos.astype(str)
            pred_y_svm_pos[pred_y_svm_pos=='0']= 'mittig'
            pred_y_svm_pos[pred_y_svm_pos=='1']= 'rechts'
            pred_y_svm_pos[pred_y_svm_pos=='2']= 'schräg'

            #save
            res = pd.DataFrame(data=pred_y_rf, columns=columns_names[0:3])
            res['Position']=pred_y_svm_pos
            res.index= range(1, len(daten) + 1)
            
            
            foldername= 'Vorhersagen_' + folder_name
            res.insert(0, 'Dateiname', filenames_predict)
            

            if os.path.isfile('data.xlsx'):
                with pd.ExcelWriter('data.xlsx', mode='a', engine='openpyxl') as writer:
                    if foldername in writer.book.sheetnames:
                        writer.book.remove(writer.book[foldername])
                    res.to_excel(writer, sheet_name=foldername, index=True)
                    workbook = writer.book
                    worksheet = writer.sheets[foldername]
                    
                    # Set the column width and format.
                    for i, col in enumerate(res.columns):
                        column_len = max(res[col].astype(str).map(len).max(), len(col)) + 2  # Add a little padding
                        worksheet.column_dimensions[worksheet.cell(1, i+1).column_letter].width = column_len
            else:
                with pd.ExcelWriter('data.xlsx', engine='xlsxwriter') as writer:
                    res.to_excel(writer, sheet_name=foldername, index=True)
                    workbook = writer.book
                    worksheet = writer.sheets[foldername]
                    
                    # Set the column width and format.
                    for i, col in enumerate(res.columns):
                        column_len = max(res[col].astype(str).map(len).max(), len(col)) + 2  # Add a little padding
                        worksheet.set_column(i, i, column_len)
            del res, pred_y_rf, pred_y_svm_pos

def find_high(pc_path, filename):
    
    high= np.zeros(3)
    high[high==0]=-1
    
    data = pd.read_csv(pc_path + '/' + filename)

    data=data.values
            
    #filter based on knowing the positionen in x and y
    lower_threshold_x=0.1
    mask_x = data[:, 0] <= lower_threshold_x 
    filtered_data = data[mask_x]

    threshold_y=0.05
    mask_y = filtered_data[:, 1] <= threshold_y
    filtered_data = filtered_data[mask_y]

            # KMeans Cluster on z values to finde ground 
    z_data = filtered_data[:,2]
    n_clusters = 2
    initial_centroids = np.array([[np.max(z_data)], [np.min(z_data)]])
            
    z_data = filtered_data[:,2].reshape(-1, 1)
    kmeans = KMeans(n_clusters=n_clusters, init=initial_centroids, n_init=1)

    kmeans.fit(z_data)

    labels = kmeans.labels_
    centroids = kmeans.cluster_centers_

    ground_offset= np.max(centroids[:][0])
    high[0]=ground_offset
    ground_label = np.argmax(centroids[:][0])
    z_part = z_data[labels != ground_label]
    ground_p = z_data[labels == ground_label]
    mask_z = filtered_data[:,2] < np.min(ground_p)
    filtered_data = filtered_data[mask_z]
     
    tmp = filtered_data.copy()
            
    mask_mod1= tmp[:,2] <= np.min(tmp[:,2]) + 0.001
    f_z =  tmp[mask_mod1]
    h1= np.mean(f_z[:,2])
    high[1] =ground_offset - h1 

    mask_del = np.all(np.isin(tmp, f_z, invert=True), axis=1)
    tmp = tmp[mask_del]
    mask_del = tmp[:,2] > h1 + 0.005
    tmp = tmp[mask_del]

    #find dorn high with dbscan and modus
    eps = 0.005  # maximum distance between two samples for one to be considered as in the neighborhood of the other
    min_samples = 40  # number of samples (or total weight) in a neighborhood for a point to be considered as a core point
    dbscan = DBSCAN(eps=eps, min_samples=min_samples)
    labels = dbscan.fit_predict(tmp)
            
    u_labels = np.unique(labels[labels != -1])
    tmp_h= np.zeros(len(u_labels))

    if len(u_labels) == 1:
        high[2] = ground_offset - np.mean(tmp[labels==0][:,2])
            
    else:
        var_x=np.zeros(len(u_labels))
        var_y=np.zeros(len(u_labels))
        var=np.zeros(len(u_labels))
        for j in range(len(u_labels)):
            var_x[j]=np.var(tmp[labels==j][:,0])
            var_y[j]=np.var(tmp[labels==j][:,1])
            var[j] = np.linalg.norm([var_x, var_y])
                        
        choosen_cluster= np.argmax(var) 
                            
        test_1 = [tuple(row) for row in filtered_data]
        test_2 = [tuple(row) for row in tmp[labels==choosen_cluster]]

        # Find indices of rows in array1 that are not in array2
        indices_to_keep = [i for i, row in enumerate(test_1) if row not in test_2]
                                    
        h3 = tmp[labels==choosen_cluster][:,2]
        max_h3= np.max(h3)
        h3 = h3[h3 > max_h3- 0.0025]
        high[2] = ground_offset - np.mean(h3)

    return high

def save_high(high_data, filenames_high):
    high_labels=['Abstand der Kamera zum Schamottstein', 'Schmiedeteilhöhe', 'Dornhöhe']
    size= len(filenames_high)
    index=np.arange(1,size+1,1)
    df = pd.DataFrame(high_data, columns=high_labels, index=index)
    foldername= 'Höhendaten'
    df.insert(0, 'Dateiname', filenames_high)
    

    if os.path.isfile('data.xlsx'):
        with pd.ExcelWriter('data.xlsx', mode='a', engine='openpyxl') as writer:
            if foldername in writer.book.sheetnames:
                writer.book.remove(writer.book[foldername])
            df.to_excel(writer, sheet_name=foldername, index=True)
            workbook = writer.book
            worksheet = writer.sheets[foldername]
                    
            # Set the column width and format.
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2  # Add a little padding
                worksheet.column_dimensions[worksheet.cell(1, i+1).column_letter].width = column_len
    else:
        with pd.ExcelWriter('data.xlsx', engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name=foldername, index=True)
            workbook = writer.book
            worksheet = writer.sheets[foldername]
                    
            # Set the column width and format.
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2  # Add a little padding
                worksheet.set_column(i, i, column_len)

def adjust_column_width(file_path):
    # Load the workbook
    wb = load_workbook(filename=file_path)
    
    # Iterate through each sheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Iterate through each column in the sheet
        for col in ws.columns:
            max_length = 0
            # Iterate through each cell in the column to find the maximum length
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Set the column width to accommodate the maximum length
            adjusted_width = (max_length + 2) * 1.2 # Additional padding
            column_letter = get_column_letter(col[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the modified workbook
    wb.save(filename=file_path)

def main(x=650, y=400):
    img_path= os.path.abspath(os.getcwd()) + '/Messung'
    save_path= os.path.abspath(os.getcwd()) + '/Messung'
    folder_path= save_path + "/Ergebnisse"
    if os.path.exists(folder_path):
        try:
            shutil.rmtree(folder_path)
        except:
            print("Excel Datei noch offen")
            sys.exit()

    # Create a new directory at the specified path
    os.makedirs(folder_path)

    folder_names = get_folder_names(img_path)
    for folder_name in folder_names:
        if folder_name != '3d' and folder_name != 'Ergebnisse':
            ip=img_path + "/" + str(folder_name)
            img_save= folder_path + "/" + folder_name
            os.makedirs(img_save)
            image_and_data(image_folder=ip  ,save_path=img_save, speichern_data=True, foldername=folder_name, x=x, y=y)
        
        if folder_name == '3d':
            high_data=[]
            filenames_high=[]
            pc_path=os.path.abspath(os.getcwd()) + '/Messung/' + folder_name
            for filename in sorted(os.listdir(pc_path), key=natural_sort_key):
                if filename.endswith('.csv') :
                    filenames_high.append(filename)
                    high_data.append(find_high(pc_path, filename))
            save_high(high_data, filenames_high)

    source_file = os.path.abspath(os.getcwd()) + '\data.xlsx'
    predection_daten_images(folder_names=folder_names, data_path=source_file)


    adjust_column_width(source_file)
     
        
        
    
    
    destination_folder = os.path.abspath(os.getcwd()) + "\Messung/Ergebnisse"
    shutil.move(source_file, destination_folder)


    

if __name__ == '__main__':
    main()